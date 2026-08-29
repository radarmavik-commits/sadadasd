import flet as ft
from google import genai
from google.genai import types
import json
import pandas as pd


APP_TITLE = "Кошторис AI"


def clean_json(text: str) -> str:
    text = (text or "").strip()
    if "```json" in text:
        text = text.split("```json", 1)[1]
    elif "```" in text:
        text = text.split("```", 1)[1]
    if "```" in text:
        text = text.split("```", 1)[0]
    return text.strip()


def safe_number(value, default=0):
    try:
        if value is None:
            return default
        if isinstance(value, str):
            value = value.replace(",", ".").replace(" ", "")
        return float(value)
    except (ValueError, TypeError):
        return default


def money(value):
    value = safe_number(value)
    if abs(value - round(value)) < 0.001:
        return f"{int(round(value)):,}".replace(",", " ") + " грн"
    return f"{value:,.2f}".replace(",", " ").replace(".", ",") + " грн"


def get_grounding_sources(response):
    sources = []
    try:
        candidates = getattr(response, "candidates", None) or []
        if not candidates:
            return sources

        metadata = getattr(candidates[0], "grounding_metadata", None)
        if not metadata:
            return sources

        chunks = getattr(metadata, "grounding_chunks", None) or []

        for chunk in chunks:
            web = getattr(chunk, "web", None)
            if not web:
                continue

            uri = getattr(web, "uri", None)
            title = getattr(web, "title", None)

            if uri:
                sources.append({
                    "title": title or uri,
                    "url": uri,
                })
    except Exception:
        pass

    unique = []
    seen = set()
    for source in sources:
        url = source["url"]
        if url not in seen:
            seen.add(url)
            unique.append(source)

    return unique


def main(page: ft.Page):
    page.title = APP_TITLE
    page.theme_mode = ft.ThemeMode.LIGHT
    page.bgcolor = "#F6F8FC"
    page.padding = 0
    page.scroll = ft.ScrollMode.AUTO

    current_sections = []
    current_sources = []
    current_city = "Київ"

    city_dropdown = ft.Dropdown(
        label="Місто / регіон",
        value="Київ",
        options=[
            ft.dropdown.Option(x)
            for x in [
                "Київ", "Дніпро", "Львів", "Одеса", "Харків",
                "Запоріжжя", "Вінниця", "Полтава", "Інше місто"
            ]
        ],
        border_radius=12,
        border_color="#D9E2F2",
        focused_border_color="#1769E0",
        text_size=14,
        content_padding=14,
    )

    works_input = ft.TextField(
        label="Список виконаних робіт",
        multiline=True,
        min_lines=7,
        max_lines=12,
        hint_text=(
            "Наприклад:\n"
            "Укладання плитки 20 м²\n"
            "Шпаклівка стін 50 м²\n"
            "Поклейка шпалер 30 м²\n"
            "Монтаж гіпсокартону 15 м²"
        ),
        border_radius=12,
        border_color="#D9E2F2",
        focused_border_color="#1769E0",
        text_size=14,
        content_padding=14,
    )

    api_key_input = ft.TextField(
        label="Google Gemini API Key",
        password=True,
        can_reveal_password=True,
        border_radius=12,
        border_color="#D9E2F2",
        focused_border_color="#1769E0",
        text_size=14,
        content_padding=14,
    )

    status_text = ft.Text("", size=13, weight=ft.FontWeight.W_500)
    results_container = ft.Column(spacing=12)
    total_container = ft.Container(visible=False)
    sources_container = ft.Column(spacing=6, visible=False)

    def show_status(text, color="#53657D"):
        status_text.value = text
        status_text.color = color
        page.update()

    def work_total(section):
        return sum(
            safe_number(item.get("work_total"))
            for item in section.get("items", [])
        )

    def grand_total(sections):
        return sum(work_total(s) for s in sections)

    def close_dialog(dialog):
        dialog.open = False
        page.update()

    def show_sources(e=None):
        if not current_sources:
            show_status(
                "ℹ️ Gemini не повернув окремих веб-джерел для цього розрахунку.",
                "#D97706",
            )
            return

        content = ft.Column(spacing=8, scroll=ft.ScrollMode.AUTO)

        for i, source in enumerate(current_sources, 1):
            title = source["title"]
            url = source["url"]

            content.controls.append(
                ft.Container(
                    bgcolor="#F7F9FC",
                    border_radius=10,
                    padding=10,
                    content=ft.Column(
                        spacing=4,
                        controls=[
                            ft.Text(
                                f"{i}. {title}",
                                size=12,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Text(
                                url,
                                size=10,
                                color="#1769E0",
                                selectable=True,
                            ),
                        ],
                    ),
                )
            )

        dialog = ft.AlertDialog(
            title=ft.Text("Джерела цін"),
            content=ft.Container(
                width=500,
                height=350,
                content=content,
            ),
            actions=[
                ft.TextButton(
                    "Закрити",
                    on_click=lambda _: close_dialog(dialog),
                )
            ],
        )

        page.dialog = dialog
        dialog.open = True
        page.update()

    def create_work_card(number, item):
        name = item.get("work_name", "Робота")
        unit = item.get("work_unit", "шт")
        qty = safe_number(item.get("work_qty"))
        avg = safe_number(item.get("avg_price"))
        min_price = safe_number(item.get("min_price"))
        max_price = safe_number(item.get("max_price"))
        total = safe_number(item.get("work_total"))

        range_text = (
            f"{money(min_price)} — {money(max_price)}"
            if min_price > 0 and max_price > 0
            else "Діапазон не визначено"
        )

        return ft.Container(
            bgcolor="#FFFFFF",
            border=ft.Border.all(1, "#E2E8F2"),
            border_radius=14,
            padding=14,
            content=ft.Column(
                spacing=8,
                controls=[
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Row(
                                spacing=10,
                                controls=[
                                    ft.Container(
                                        width=34,
                                        height=34,
                                        border_radius=10,
                                        bgcolor="#EEF4FF",
                                        alignment=ft.alignment.center,
                                        content=ft.Text(
                                            str(number),
                                            color="#1769E0",
                                            weight=ft.FontWeight.BOLD,
                                        ),
                                    ),
                                    ft.Column(
                                        spacing=2,
                                        controls=[
                                            ft.Text(
                                                name,
                                                size=14,
                                                weight=ft.FontWeight.BOLD,
                                                color="#172B4D",
                                            ),
                                            ft.Text(
                                                f"{qty:g} {unit}",
                                                size=12,
                                                color="#718096",
                                            ),
                                        ],
                                    ),
                                ],
                            ),
                            ft.Text(
                                money(total),
                                size=14,
                                weight=ft.FontWeight.BOLD,
                                color="#168A5B",
                            ),
                        ],
                    ),
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text("Середня ціна", size=12, color="#718096"),
                            ft.Text(
                                f"{money(avg)} / {unit}",
                                size=12,
                                weight=ft.FontWeight.W_500,
                                color="#263A5A",
                            ),
                        ],
                    ),
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text("Знайдений діапазон", size=12, color="#718096"),
                            ft.Text(
                                range_text,
                                size=12,
                                color="#52627A",
                            ),
                        ],
                    ),
                ],
            ),
        )

    def create_section_card(section):
        name = section.get("section_name", "РОЗДІЛ")
        items = section.get("items", [])
        total = work_total(section)

        return ft.Container(
            bgcolor="#FFFFFF",
            border=ft.Border.all(1, "#E0E7F0"),
            border_radius=16,
            padding=16,
            content=ft.Column(
                spacing=10,
                controls=[
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Row(
                                spacing=10,
                                controls=[
                                    ft.Container(
                                        width=40,
                                        height=40,
                                        border_radius=12,
                                        bgcolor="#F0E9FF",
                                        alignment=ft.alignment.center,
                                        content=ft.Text(
                                            "▦", size=20, color="#7546D8"
                                        ),
                                    ),
                                    ft.Column(
                                        spacing=2,
                                        controls=[
                                            ft.Text(
                                                name,
                                                size=15,
                                                weight=ft.FontWeight.BOLD,
                                                color="#172B4D",
                                            ),
                                            ft.Text(
                                                f"{len(items)} позицій",
                                                size=12,
                                                color="#7B879A",
                                            ),
                                        ],
                                    ),
                                ],
                            ),
                            ft.Text(
                                money(total),
                                size=14,
                                weight=ft.FontWeight.BOLD,
                                color="#1769E0",
                            ),
                        ],
                    ),
                    ft.Divider(height=4, color="#EDF1F7"),
                    *[
                        create_work_card(i, item)
                        for i, item in enumerate(items, 1)
                    ],
                    ft.Container(
                        bgcolor="#F4F8FF",
                        border_radius=10,
                        padding=10,
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            controls=[
                                ft.Text(
                                    "Разом по розділу",
                                    size=12,
                                    color="#52627A",
                                ),
                                ft.Text(
                                    money(total),
                                    size=13,
                                    weight=ft.FontWeight.BOLD,
                                    color="#1769E0",
                                ),
                            ],
                        ),
                    ),
                ],
            ),
        )

    def render_results(sections):
        results_container.controls.clear()

        total = grand_total(sections)
        count = sum(len(s.get("items", [])) for s in sections)

        results_container.controls.append(
            ft.Row(
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                controls=[
                    ft.Text(
                        "Результат кошторису",
                        size=20,
                        weight=ft.FontWeight.BOLD,
                        color="#172B4D",
                    ),
                    ft.Text(
                        f"{count} позицій",
                        size=12,
                        color="#7B879A",
                    ),
                ],
            )
        )

        for section in sections:
            results_container.controls.append(create_section_card(section))

        total_container.content = ft.Container(
            bgcolor="#1769E0",
            border_radius=18,
            padding=18,
            content=ft.Column(
                spacing=7,
                controls=[
                    ft.Text(
                        "ЗАГАЛЬНА ВАРТІСТЬ РОБІТ",
                        size=12,
                        color="#DCE9FF",
                        weight=ft.FontWeight.BOLD,
                    ),
                    ft.Text(
                        money(total),
                        size=28,
                        color="#FFFFFF",
                        weight=ft.FontWeight.BOLD,
                    ),
                    ft.Text(
                        "Матеріали не враховуються",
                        size=12,
                        color="#DCE9FF",
                    ),
                ],
            ),
        )
        total_container.visible = True

        sources_container.controls.clear()

        if current_sources:
            sources_container.controls.append(
                ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[
                        ft.Text(
                            f"Використано веб-джерел: {len(current_sources)}",
                            size=12,
                            color="#718096",
                        ),
                        ft.TextButton(
                            "Переглянути джерела",
                            on_click=show_sources,
                        ),
                    ],
                )
            )
            sources_container.visible = True
        else:
            sources_container.visible = False

    def generate_excel(sections, sources):
        rows = []
        counter = 1

        for section in sections:
            section_name = section.get("section_name", "РОЗДІЛ")

            rows.append({
                "№": "",
                "Розділ": section_name,
                "Найменування роботи": "",
                "Од. вим.": "",
                "Кількість": "",
                "Мін. ціна, грн": "",
                "Середня ціна, грн": "",
                "Макс. ціна, грн": "",
                "Вартість, грн": "",
            })

            for item in section.get("items", []):
                rows.append({
                    "№": counter,
                    "Розділ": section_name,
                    "Найменування роботи": item.get("work_name", ""),
                    "Од. вим.": item.get("work_unit", ""),
                    "Кількість": safe_number(item.get("work_qty")),
                    "Мін. ціна, грн": safe_number(item.get("min_price")),
                    "Середня ціна, грн": safe_number(item.get("avg_price")),
                    "Макс. ціна, грн": safe_number(item.get("max_price")),
                    "Вартість, грн": safe_number(item.get("work_total")),
                })
                counter += 1

        df = pd.DataFrame(rows)
        path = "koshtorys_ai.xlsx"

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Кошторис")

            if sources:
                source_df = pd.DataFrame(sources)
                source_df.to_excel(
                    writer,
                    index=False,
                    sheet_name="Джерела",
                )

        try:
            from openpyxl import load_workbook

            wb = load_workbook(path)

            ws = wb["Кошторис"]
            widths = {
                "A": 8, "B": 28, "C": 42, "D": 14, "E": 14,
                "F": 18, "G": 20, "H": 18, "I": 20,
            }

            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)

            if "Джерела" in wb.sheetnames:
                src = wb["Джерела"]
                src.column_dimensions["A"].width = 55
                src.column_dimensions["B"].width = 80

            wb.save(path)
        except Exception:
            pass

        return path

    def export_excel(e):
        if not current_sections:
            show_status("Спочатку сформуйте кошторис.", "#D97706")
            return

        try:
            path = generate_excel(current_sections, current_sources)
            show_status(f"✅ Excel збережено: {path}", "#168A5B")
        except Exception as ex:
            show_status(f"❌ Не вдалося створити Excel: {ex}", "#C62828")

    def calculate_click(e):
        nonlocal current_sections, current_sources, current_city

        api_key = api_key_input.value.strip()
        raw_text = works_input.value.strip()
        city = city_dropdown.value or "Київ"

        current_city = city

        if not api_key:
            show_status("⚠️ Введіть Gemini API Key.", "#C62828")
            return

        if not raw_text:
            show_status("⚠️ Введіть список виконаних робіт.", "#C62828")
            return

        show_status(
            "⏳ Gemini шукає актуальні ціни в інтернеті...",
            "#1769E0",
        )

        results_container.controls.clear()
        total_container.visible = False
        sources_container.visible = False
        page.update()

        try:
            client = genai.Client(api_key=api_key)

            grounding_tool = types.Tool(
                google_search=types.GoogleSearch()
            )

            prompt = f"""
Ти професійний кошторисник будівельних та ремонтних робіт в Україні.

Місто/регіон: {city}
Поточний рік: 2026.

Користувач надав список ВИКОНАНИХ РОБІТ.
Тобі потрібно оцінити актуальну середню РИНКОВУ ЦІНУ САМЕ ЗА РОБОТУ,
без вартості матеріалів.

ОБОВ'ЯЗКОВО:
1. Використовуй Google Search для пошуку актуальних пропозицій.
2. Шукай ціни саме для міста/регіону {city}, якщо це можливо.
3. Перевагу надавай українським сайтам, прайс-листам будівельних компаній,
   сервісам пошуку майстрів та локальним пропозиціям.
4. Не включай матеріали, доставку, інструмент або товар у ціну роботи.
5. Якщо є кілька цін, визнач орієнтовний мінімум, максимум та середню ціну.
6. Відкидай очевидні аномальні значення.
7. Якщо локальних даних мало, використовуй дані по Україні,
   але враховуй це в полі "price_note".
8. Не вигадуй кількість. Якщо кількість є у тексті користувача —
   використовуй її.
9. Розбий роботи на логічні розділи.
10. work_total = work_qty * avg_price.
11. Матеріали НЕ ПОВЕРТАЙ.

Список користувача:
{raw_text}

Поверни ТІЛЬКИ валідний JSON такого формату:

{{
  "sections": [
    {{
      "section_name": "ОЗДОБЛЮВАЛЬНІ РОБОТИ",
      "items": [
        {{
          "work_name": "Укладання плитки",
          "work_unit": "м²",
          "work_qty": 20,
          "min_price": 500,
          "avg_price": 650,
          "max_price": 800,
          "work_total": 13000,
          "price_note": "Ціна за роботу без матеріалів"
        }}
      ]
    }}
  ]
}}

Тільки JSON без markdown.
"""

            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt,
                config=types.GenerateContentConfig(
                    temperature=0.1,
                    tools=[grounding_tool],
                ),
            )

            current_sources = get_grounding_sources(response)

            data = json.loads(clean_json(response.text))
            sections = data.get("sections", [])

            normalized = []

            for section in sections:
                items = []

                for item in section.get("items", []):
                    qty = max(0, safe_number(item.get("work_qty")))
                    avg = max(0, safe_number(item.get("avg_price")))
                    min_price = max(0, safe_number(item.get("min_price")))
                    max_price = max(0, safe_number(item.get("max_price")))

                    if min_price and max_price and min_price > max_price:
                        min_price, max_price = max_price, min_price

                    if not min_price:
                        min_price = avg

                    if not max_price:
                        max_price = avg

                    items.append({
                        "work_name": str(
                            item.get("work_name", "Робота")
                        ),
                        "work_unit": str(
                            item.get("work_unit", "шт")
                        ),
                        "work_qty": qty,
                        "min_price": min_price,
                        "avg_price": avg,
                        "max_price": max_price,
                        "work_total": round(qty * avg, 2),
                        "price_note": str(
                            item.get(
                                "price_note",
                                "Орієнтовна ринкова ціна за роботу",
                            )
                        ),
                    })

                if items:
                    normalized.append({
                        "section_name": str(
                            section.get(
                                "section_name",
                                "ІНШІ РОБОТИ",
                            )
                        ),
                        "items": items,
                    })

            if not normalized:
                show_status(
                    "⚠️ AI не знайшов виконаних робіт.",
                    "#D97706",
                )
                return

            current_sections = normalized
            render_results(current_sections)

            if current_sources:
                show_status(
                    f"✅ Кошторис сформовано. Gemini використав "
                    f"{len(current_sources)} веб-джерел.",
                    "#168A5B",
                )
            else:
                show_status(
                    "⚠️ Кошторис сформовано, але джерела пошуку не повернулися.",
                    "#D97706",
                )

        except json.JSONDecodeError:
            show_status(
                "❌ Gemini повернув некоректний JSON. Спробуйте ще раз.",
                "#C62828",
            )
        except Exception as ex:
            show_status(f"❌ Помилка: {ex}", "#C62828")

        page.update()

    # Інтерфейс
    header = ft.Container(
        bgcolor="#FFFFFF",
        padding=ft.Padding.only(left=20, right=20, top=18, bottom=16),
        border=ft.Border.only(bottom=ft.BorderSide(1, "#E6EBF2")),
        content=ft.Row(
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            controls=[
                ft.Text(APP_TITLE, size=22, weight=ft.FontWeight.BOLD, color="#172B4D"),
            ],
        ),
    )

    intro = ft.Container(
        padding=ft.Padding.only(left=20, right=20, top=18),
        content=ft.Column(
            spacing=4,
            controls=[
                ft.Text("Розрахунок вартості будівельних робіт", size=16, weight=ft.FontWeight.BOLD, color="#172B4D"),
                ft.Text("Введіть список робіт та отримайте орієнтовну кошторисну вартість", size=13, color="#718096"),
            ],
        ),
    )

    calc_button = ft.ElevatedButton(
        "Розрахувати кошторис",
        icon=ft.Icons.CALCULATE,
        bgcolor="#1769E0",
        color="#FFFFFF",
        style=ft.ButtonStyle(
            shape=ft.RoundedRectangleBorder(radius=12),
            padding=18,
        ),
        on_click=calculate_click,
    )

    export_button = ft.OutlinedButton(
        "Завантажити в Excel",
        icon=ft.Icons.DOWNLOAD,
        style=ft.ButtonStyle(
            shape=ft.RoundedRectangleBorder(radius=12),
            padding=18,
        ),
        on_click=export_excel,
    )

    export_row = ft.Row(
        spacing=12,
        controls=[calc_button, export_button],
    )

    form_container = ft.Container(
        padding=ft.Padding.only(left=20, right=20),
        content=ft.Column(
            spacing=12,
            controls=[
                api_key_input,
                city_dropdown,
                works_input,
                status_text,
            ],
        ),
    )

    page.add(
        ft.Column(
            spacing=14,
            controls=[
                header,
                intro,
                form_container,
                ft.Container(
                    padding=ft.Padding.only(left=20, right=20),
                    content=results_container,
                ),
                ft.Container(
                    padding=ft.Padding.only(left=20, right=20),
                    content=total_container,
                ),
                ft.Container(
                    padding=ft.Padding.only(left=20, right=20),
                    content=sources_container,
                ),
                ft.Container(
                    padding=ft.Padding.only(left=20, right=20, bottom=24),
                    content=export_row,
                ),
            ],
        )
    )


if __name__ == "__main__":
    ft.run(main)